#! -*- coding: utf-8 -*-
from typing import Iterable
from openpyxl.cell import Cell
from openpyxl import load_workbook
import json
import sqlite3

workbook = load_workbook('Codes.xlsx',  data_only=True)

sheet = workbook['ОКПД2 - ТН ВЭД']

# cells = sheet['A1': 'D5836']

data_okpd2 = {}
data_tnved = {}


current_okpd2 = None
current_tnved = None

matc = []
tnved_new = []
okpd2_new = []


for row in sheet.iter_rows(min_row=5):
    okpd2, okpd2_desc, tnved, tnved_desc = row
    okpd2_value = okpd2.value
    tnved_value = tnved.value
    
    if okpd2_value and okpd2_value != current_okpd2:
        okpd2_value = okpd2_value.rstrip()
        current_okpd2 = okpd2_value
        data_okpd2.update({'code': okpd2_value, 'description': okpd2_desc.value})
        # okpd2_new.append({'code': okpd2_value, 'description': okpd2_desc.value})
        
        
    if tnved_value and tnved_value != current_tnved:
        tnved_value = tnved_value.rstrip()
        current_tnved = tnved_value
        data_tnved.update({'code': tnved_value, 'description': tnved_desc.value})
        # tnved_new.append({'code': tnved_value, 'description': tnved_desc.value})
    
    if all((tnved_value, okpd2_value)):
        matc.append({'tnved': tnved_value, 'okpd2': okpd2_value})

    if not tnved_value:
        matc.append({'okpd2': okpd2_value, 'tnved': current_tnved})

    if not okpd2_value:
        matc.append({'okpd2': current_okpd2, 'tnved': tnved_value})
    


def save_json(filename: str, array: Iterable):
    with open(filename, 'w') as f:
        f.write(json.dumps(array))


# print(tnved_new)
save_json('matc.json', matc)
# save_json('tnved_new.json', tnved_new)
# save_json('okpd2_new.json', okpd2_new)